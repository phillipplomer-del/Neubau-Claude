#!/usr/bin/env python3
"""
Excel File Analyzer for ERP Data
Analyzes three Excel files and extracts column structure and sample data
"""

import sys
import os

# Try to import required libraries
try:
    import pandas as pd
    use_pandas = True
    print("Using pandas for analysis")
except ImportError:
    use_pandas = False
    print("Pandas not available, trying openpyxl...")

if not use_pandas:
    try:
        from openpyxl import load_workbook
        use_openpyxl = True
        print("Using openpyxl for analysis")
    except ImportError:
        print("ERROR: Neither pandas nor openpyxl is available!")
        print("Please install one of them:")
        print("  pip install pandas openpyxl")
        print("  or")
        print("  pip install openpyxl")
        sys.exit(1)

# File paths
FILES = {
    'Sales (Offene Lieferungen)': '/Users/phillipplomer/Desktop/Programme/Neubau Claude/2025-09-30_Offene_Lieferungen_Stand.xlsx',
    'Production Planning': '/Users/phillipplomer/Desktop/Programme/Neubau Claude/2025-12-10_PP_SollIstVergleich.xlsm',
    'Project Management (Controlling)': '/Users/phillipplomer/Desktop/Programme/Neubau Claude/Controlling.xlsx'
}


def analyze_with_pandas(filepath, name):
    """Analyze Excel file using pandas"""
    print(f"\n{'='*100}")
    print(f"FILE: {name}")
    print(f"Path: {filepath}")
    print('='*100)

    try:
        # Read Excel file
        df = pd.read_excel(filepath, engine='openpyxl')

        # Basic information
        print(f"\nBASIC INFORMATION:")
        print(f"  Total rows (including header): {len(df) + 1}")
        print(f"  Data rows: {len(df)}")
        print(f"  Total columns: {len(df.columns)}")

        # All column names
        print(f"\nALL COLUMN NAMES ({len(df.columns)} columns):")
        for i, col in enumerate(df.columns, 1):
            print(f"  {i:2d}. {col}")

        # Sample data - first 5 rows
        print(f"\nSAMPLE DATA (First 5 rows):")
        print("-" * 100)

        # Show data in a readable format
        for idx in range(min(5, len(df))):
            print(f"\nRow {idx + 1}:")
            for col in df.columns:
                value = df.iloc[idx][col]
                if pd.notna(value):  # Only show non-null values
                    # Format value
                    if isinstance(value, float):
                        value_str = f"{value:.2f}" if value != int(value) else f"{int(value)}"
                    else:
                        value_str = str(value)

                    # Truncate long values
                    if len(value_str) > 60:
                        value_str = value_str[:60] + "..."

                    print(f"  {col}: {value_str}")

        # Column Analysis
        print(f"\n\nCOLUMN ANALYSIS:")
        print("-" * 100)

        # Article Number columns
        article_candidates = [col for col in df.columns
                            if any(x in str(col).lower() for x in ['art', 'item', 'artikel', 'material', 'nummer'])]
        if article_candidates:
            print(f"\nPotential ARTICLE NUMBER columns:")
            for col in article_candidates:
                sample_val = df[col].dropna().iloc[0] if len(df[col].dropna()) > 0 else "N/A"
                print(f"  - {col}")
                print(f"    Sample value: {sample_val}")
                print(f"    Unique values: {df[col].nunique()}")

        # Project Number columns
        project_candidates = [col for col in df.columns
                            if any(x in str(col).lower() for x in ['proj', 'project', 'auftrag', 'order'])]
        if project_candidates:
            print(f"\nPotential PROJECT NUMBER columns:")
            for col in project_candidates:
                sample_val = df[col].dropna().iloc[0] if len(df[col].dropna()) > 0 else "N/A"
                print(f"  - {col}")
                print(f"    Sample value: {sample_val}")
                print(f"    Unique values: {df[col].nunique()}")

        # Quantity columns
        qty_candidates = [col for col in df.columns
                        if any(x in str(col).lower() for x in ['qty', 'quantity', 'menge', 'anzahl', 'rem'])]
        if qty_candidates:
            print(f"\nPotential QUANTITY columns:")
            for col in qty_candidates:
                if pd.api.types.is_numeric_dtype(df[col]):
                    print(f"  - {col}")
                    print(f"    Min: {df[col].min()}, Max: {df[col].max()}, Mean: {df[col].mean():.2f}")

        # Date columns
        date_candidates = [col for col in df.columns
                         if any(x in str(col).lower() for x in ['date', 'datum', 'termin', 'deadline'])]
        if date_candidates:
            print(f"\nPotential DATE columns:")
            for col in date_candidates:
                sample_val = df[col].dropna().iloc[0] if len(df[col].dropna()) > 0 else "N/A"
                print(f"  - {col}")
                print(f"    Sample value: {sample_val}")

        # Status columns
        status_candidates = [col for col in df.columns
                           if any(x in str(col).lower() for x in ['status', 'state', 'zustand'])]
        if status_candidates:
            print(f"\nPotential STATUS columns:")
            for col in status_candidates:
                unique_vals = df[col].unique()[:5]  # First 5 unique values
                print(f"  - {col}")
                print(f"    Unique values (first 5): {unique_vals}")

        # Special handling for Sales file
        if 'QuantityRem1' in df.columns:
            print(f"\n{'*'*100}")
            print(f"IMPORTANT: QuantityRem1 column found (BUSINESS RULE)")
            print(f"{'*'*100}")
            zero_count = len(df[df['QuantityRem1'] == 0])
            non_zero_count = len(df[df['QuantityRem1'] > 0])
            total = len(df)
            print(f"  Total rows: {total}")
            print(f"  Rows where QuantityRem1 = 0 (already delivered, SKIP): {zero_count} ({zero_count/total*100:.1f}%)")
            print(f"  Rows where QuantityRem1 > 0 (pending delivery, INCLUDE): {non_zero_count} ({non_zero_count/total*100:.1f}%)")
            print(f"\n  ACTION: Filter out rows where QuantityRem1 == 0 before processing")

        # Recommendations
        print(f"\n\nRECOMMENDATIONS:")
        print("-" * 100)

        if article_candidates:
            print(f"  ARTIKELNUMMER (Article Number): Use column '{article_candidates[0]}'")
        else:
            print(f"  ARTIKELNUMMER (Article Number): Not clearly identified - manual review needed")

        if project_candidates:
            print(f"  PROJEKTNUMMER (Project Number): Use column '{project_candidates[0]}'")
        else:
            print(f"  PROJEKTNUMMER (Project Number): Not clearly identified - manual review needed")

    except Exception as e:
        print(f"\nERROR analyzing file: {str(e)}")
        import traceback
        traceback.print_exc()


def analyze_with_openpyxl(filepath, name):
    """Analyze Excel file using openpyxl (fallback method)"""
    print(f"\n{'='*100}")
    print(f"FILE: {name}")
    print(f"Path: {filepath}")
    print('='*100)

    try:
        # Load workbook
        wb = load_workbook(filepath, read_only=True, data_only=True)
        sheet = wb.active

        print(f"\nBASIC INFORMATION:")
        print(f"  Sheet name: {sheet.title}")
        print(f"  Total rows: {sheet.max_row}")
        print(f"  Total columns: {sheet.max_column}")

        # Get column names
        column_names = []
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            column_names.append(cell_value if cell_value is not None else f"Column_{col}")

        # Display column names
        print(f"\nALL COLUMN NAMES ({len(column_names)} columns):")
        for i, col in enumerate(column_names, 1):
            print(f"  {i:2d}. {col}")

        # Display sample data
        print(f"\nSAMPLE DATA (First 5 rows):")
        print("-" * 100)

        max_rows = min(6, sheet.max_row)
        for row in range(2, max_rows + 1):
            print(f"\nRow {row - 1}:")
            for col_idx, col_name in enumerate(column_names, 1):
                cell_value = sheet.cell(row=row, column=col_idx).value
                if cell_value is not None:
                    value_str = str(cell_value)
                    if len(value_str) > 60:
                        value_str = value_str[:60] + "..."
                    print(f"  {col_name}: {value_str}")

        # Column analysis
        print(f"\n\nCOLUMN ANALYSIS:")
        print("-" * 100)

        article_candidates = [col for col in column_names
                            if col and any(x in str(col).lower() for x in ['art', 'item', 'artikel', 'material'])]
        project_candidates = [col for col in column_names
                            if col and any(x in str(col).lower() for x in ['proj', 'project', 'auftrag'])]
        qty_candidates = [col for col in column_names
                        if col and any(x in str(col).lower() for x in ['qty', 'quantity', 'menge', 'anzahl', 'rem'])]
        date_candidates = [col for col in column_names
                         if col and any(x in str(col).lower() for x in ['date', 'datum', 'termin'])]

        if article_candidates:
            print(f"\nPotential ARTICLE NUMBER columns: {article_candidates}")
        if project_candidates:
            print(f"\nPotential PROJECT NUMBER columns: {project_candidates}")
        if qty_candidates:
            print(f"\nPotential QUANTITY columns: {qty_candidates}")
        if date_candidates:
            print(f"\nPotential DATE columns: {date_candidates}")

        # Check for QuantityRem1
        if 'QuantityRem1' in column_names:
            print(f"\n{'*'*100}")
            print(f"IMPORTANT: QuantityRem1 column found")
            print(f"  BUSINESS RULE: Rows with QuantityRem1 = 0 should be skipped (already delivered)")
            print(f"{'*'*100}")

        wb.close()

    except Exception as e:
        print(f"\nERROR analyzing file: {str(e)}")
        import traceback
        traceback.print_exc()


def main():
    """Main analysis function"""
    print("="*100)
    print("EXCEL FILE ANALYZER - ERP Data Analysis")
    print("="*100)

    # Check files exist
    print("\nChecking files...")
    for name, filepath in FILES.items():
        exists = "EXISTS" if os.path.exists(filepath) else "NOT FOUND"
        print(f"  {name}: {exists}")

    print("\n")

    # Analyze each file
    for name, filepath in FILES.items():
        if not os.path.exists(filepath):
            print(f"\nSKIPPING {name}: File not found")
            continue

        if use_pandas:
            analyze_with_pandas(filepath, name)
        else:
            analyze_with_openpyxl(filepath, name)

    print("\n" + "="*100)
    print("ANALYSIS COMPLETE")
    print("="*100)


if __name__ == "__main__":
    main()
