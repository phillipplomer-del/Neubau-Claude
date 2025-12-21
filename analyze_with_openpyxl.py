#!/usr/bin/env python3
import openpyxl
from openpyxl import load_workbook
import os

# File paths
files = {
    'Sales (Offene Lieferungen)': '/Users/phillipplomer/Desktop/Programme/Neubau Claude/2025-09-30_Offene_Lieferungen_Stand.xlsx',
    'Production Planning': '/Users/phillipplomer/Desktop/Programme/Neubau Claude/2025-12-10_PP_SollIstVergleich.xlsm',
    'Project Management (Controlling)': '/Users/phillipplomer/Desktop/Programme/Neubau Claude/Controlling.xlsx'
}

def analyze_excel_file(filepath, name):
    print(f"\n{'='*80}")
    print(f"FILE: {name}")
    print(f"Path: {filepath}")
    print('='*80)

    try:
        # Check if file exists
        if not os.path.exists(filepath):
            print(f"ERROR: File does not exist!")
            return

        # Load workbook
        wb = load_workbook(filepath, read_only=True, data_only=True)

        # Get the active sheet (or first sheet)
        sheet = wb.active
        print(f"\nActive sheet name: {sheet.title}")
        print(f"Total rows: {sheet.max_row}")
        print(f"Total columns: {sheet.max_column}")

        # Get column names from first row
        column_names = []
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            column_names.append(cell_value if cell_value is not None else f"Column_{col}")

        # Display all column names
        print("\nALL COLUMN NAMES:")
        for i, col in enumerate(column_names, 1):
            print(f"  {i}. {col}")

        # Display first 5 data rows (rows 2-6)
        print("\nFIRST 5 DATA ROWS (SAMPLE):")
        max_rows_to_show = min(6, sheet.max_row)
        for row in range(2, max_rows_to_show + 1):
            print(f"\n  Row {row-1}:")
            for col_idx, col_name in enumerate(column_names, 1):
                cell_value = sheet.cell(row=row, column=col_idx).value
                if cell_value is not None:
                    # Truncate long values
                    value_str = str(cell_value)
                    if len(value_str) > 50:
                        value_str = value_str[:50] + "..."
                    print(f"    {col_name}: {value_str}")

        # Column analysis
        print("\n\nCOLUMN ANALYSIS:")

        # Look for Article Number columns
        article_candidates = [col for col in column_names if col and any(x in str(col).lower() for x in ['art', 'item', 'artikel', 'material'])]
        if article_candidates:
            print(f"  Potential Article Number columns: {article_candidates}")

        # Look for Project Number columns
        project_candidates = [col for col in column_names if col and any(x in str(col).lower() for x in ['proj', 'project', 'auftrag'])]
        if project_candidates:
            print(f"  Potential Project Number columns: {project_candidates}")

        # Look for quantity columns
        qty_candidates = [col for col in column_names if col and any(x in str(col).lower() for x in ['qty', 'quantity', 'menge', 'anzahl', 'rem'])]
        if qty_candidates:
            print(f"  Potential Quantity columns: {qty_candidates}")

        # Look for date columns
        date_candidates = [col for col in column_names if col and any(x in str(col).lower() for x in ['date', 'datum', 'termin'])]
        if date_candidates:
            print(f"  Potential Date columns: {date_candidates}")

        # Look for status columns
        status_candidates = [col for col in column_names if col and any(x in str(col).lower() for x in ['status', 'state', 'zustand'])]
        if status_candidates:
            print(f"  Potential Status columns: {status_candidates}")

        # Special check for QuantityRem1
        if 'QuantityRem1' in column_names:
            qty_rem_col = column_names.index('QuantityRem1') + 1
            zero_count = 0
            non_zero_count = 0

            for row in range(2, min(1000, sheet.max_row + 1)):  # Check first 1000 rows
                cell_value = sheet.cell(row=row, column=qty_rem_col).value
                if cell_value == 0:
                    zero_count += 1
                elif cell_value and cell_value > 0:
                    non_zero_count += 1

            print(f"\n  *** IMPORTANT: Found 'QuantityRem1' column ***")
            print(f"      - Rows where QuantityRem1 = 0 (sampled): {zero_count}")
            print(f"      - Rows where QuantityRem1 > 0 (sampled): {non_zero_count}")
            print(f"      - Rule: Rows with QuantityRem1 = 0 should be skipped (already delivered)")

        wb.close()

    except Exception as e:
        print(f"\nERROR reading file: {str(e)}")
        import traceback
        traceback.print_exc()

# Analyze all files
for name, filepath in files.items():
    analyze_excel_file(filepath, name)

print("\n" + "="*80)
print("ANALYSIS COMPLETE")
print("="*80)
